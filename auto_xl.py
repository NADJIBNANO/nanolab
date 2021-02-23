import openpyxl as xl
import pandas as  pd
import numpy as np 

wb = xl.load_workbook('CU0301_9998.xlsx')
sheet = wb['LMD']
for row in range(5,sheet.max_row +1):
    cell_mat = sheet.cell(row,2)
    cell_année = sheet.cell(row,3)
    cell_nom = (sheet.cell(row,4))
    cell_prenom = (sheet.cell(row,5))
   # cell_nom_prenom = cell_nom + cell_prenom

    #print(cell_nom.value)
#wb.save('new.xlsx')

com_fis = pd.read_csv('Comptabilité et fiscalité_S5.csv')
#com_fis


