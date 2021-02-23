import os
import sys
from PyPDF2 import PdfFileReader,PdfFileWriter
from openpyxl import load_workbook
#pdf_document = "thinkpython.pdf"
args = sys.argv
if len(args) > 1:
    pdf_file_path = args[1]
    print("The File Name Is:" , pdf_file_path)

#os.makedirs(str(pdf_file_path))

'''with open(pdf_file_path, "rb") as filehandle:
    pdf = PdfFileReader(filehandle)
    info = pdf.getDocumentInfo()
    pages = pdf.getNumPages()

    print (info)
    print ("number of pages: %i" % pages)

    page1 = pdf.getPage(0)
    print(page1)
    print(page1.extractText())'''


#pdf_file_path = 'thinkpython.pdf'
file_base_name = pdf_file_path.replace('.pdf', '')
output_folder_path = os.path.join(os.getcwd(), 'Output')

wb = load_workbook('Class_momb.xlsx')
sheet = wb['Feuil1']
max_row = sheet.max_row
list_nom=[]
for i in range(1, 29):
    cell_name = sheet.cell(row=i, column=1)
    nam_lastnam=str(cell_name.value)
    list_nom.append(nam_lastnam)
print(list_nom[2])

pdf = PdfFileReader(pdf_file_path)

for page_num in range(pdf.numPages):
    pdfWriter = PdfFileWriter()
    pdfWriter.addPage(pdf.getPage(page_num))
    
    with open(os.path.join(output_folder_path, '{0}.pdf'.format(list_nom[page_num+1] )), 'wb') as f:
        pdfWriter.write(f)
        print("The Splite Is Done", page_num+1, list_nom[page_num+1])
        f.close()
